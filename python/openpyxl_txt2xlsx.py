#!/usr/bin/python
#coding=utf-8
#filename: openpyxl_txt2xlsx.py
#保存IIS的日志文件（文本）为XLS表格
#为了解决(使用xlwt/xlrd限制XLS表格的行不能走过65535行)
#date: 20130726-1300
#auth: chengl6500
"""以下数字作为列宽标尺作用
12345678901234567890123456789012345678901234567890123456789012345678901234567890
"""
"""测试结果与xls格式（xlwr_txt2xls.py）对比如下
----------------------lineMax=65535-----
time is Mon Jul 29 18:28:17 2013, beginning......

end@ Mon Jul 29 18:32:36 2013
save@ Mon Jul 29 18:29:17 2013
start@ Mon Jul 29 18:28:17 2013
inst1.save_book()保存XLSX文件时很慢?????????????
xlsx为2007格式，同样包括65535条记录，文件大小为3.1M
而xls为2003格式，包括65535条记录时文件大小为14M
？？？？？？？？再试试不同记录数时，保存时间和文件大小差异？？？？
"""

import os
import sys
import time
import openpyxl

f_txt2xls = '/tmp/txt2xls.xls'
f_txt_from = '/home/love/scxhedu_ex130706.log'

class WriteXls2007():
    txt_file = None
    xls_file = None
    xls_book = None
    xls_book_sheet = None
    txt_from = None
    lineNo = 0

    def __init__(self, xls_file=None, txt_file=None, sheet_name=None):
        if not xls_file or not txt_file:
            return

        if not os.path.exists(txt_file):
            return

        if os.path.exists(xls_file):
            xls_file_bak = xls_file + time.strftime("%Y%m%d_%H%M%S")
            os.rename(xls_file, xls_file_bak)

        self.txt_file = txt_file
        self.xls_file = xls_file
        self.xls_book = openpyxl.Workbook()
        if sheet_name:
            self.xls_book_sheet = self.xls_book.worksheets[0]
            self.xls_book_sheet.title = sheet_name
        else:
            exit(-1)

    def open_txt(self, txt_file=None):
        if not txt_file:
            if self.txt_file:
                txt_file =self.txt_file
            else:
                return

        try:
            txt_from = open(txt_file)
        except IOError as e:
            print "open file %s ERROR!" %(txt_from), e
            return
        #debug#print txt_from
        self.txt_from = txt_from
        return

    def create_sheet(self, xls_book=None, sheet_name=None):
        if not xls_book:
            xls_book = self.xls_book

        if not sheet_name:
            xls_book.create_sheet()
            return

        xls_book.create_sheet(sheet_name)

        # if xls_book:
        #     if sheet_name:
        #         print xls_book
        #         xls_book.create_sheet(sheet_name)
        #         return
        #     else:
        #         xls_book.create_sheet()
        #         return

        # if not sheet_name:
        #     self.xls_book.create_sheet()
        # else:
        #     self.xls_book.create_sheet(title=sheet_name)

    def write_sheet_col(self, txt_from=None, sheet_to=None, lineNo=None):
        col = 0
        # txt_cont =  txt_from.split()
        # #print txt_cont
        # for i in xrange(len(txt_cont)):
        #     sheet_to.cell(row=lineNo, column=col).value = txt_cont[col]
        #     col += 1
        for tmp in txt_from.split():
            sheet_to.cell(row=lineNo, column=col).value = tmp
            col += 1

    def write_sheet(self, txt_from=None, sheet_to=None, lineMax=65535):
        if not txt_from:
            txt_from = self.txt_from
        if not sheet_to:
            sheet_to = self.xls_book_sheet

        txt_from_iter = iter(txt_from)
        while True:
            try:
                txt_proc = txt_from_iter.next().decode('utf-8', 'ignore')
            except StopIteration:
                break
            self.write_sheet_col(txt_proc, sheet_to, self.lineNo)
            self.lineNo += 1
            if self.lineNo == lineMax:
                break
            #debug#print lineNo

    def save_book(self, xls_book=None, file_name=None):
        if not xls_book:
            xls_book = self.xls_book
        if not file_name:
            file_name = self.xls_file

        xls_book.save(file_name)

if __name__ == "__main__":
    time_start = time.ctime()
    print "time is %s, beginning......\n" % time_start
    inst1 = WriteXls2007('/tmp/xlsx1.xlsx', '/home/love/scxhedu_ex130706.log', 'sheet2013')
    #inst1 = WriteXls2007('/tmp/xlsx1.xlsx', '/home/love/scxhedu_ex130706.log')

    #inst1.create_sheet(sheet_name='201300')
    inst1.open_txt()
    inst1.write_sheet(lineMax=65535)
    time_save = time.ctime()
    #inst1.save_book()
    time_end = time.ctime()
    print "end@", time_end, '\nsave@', time_save, '\nstart@', time_start


